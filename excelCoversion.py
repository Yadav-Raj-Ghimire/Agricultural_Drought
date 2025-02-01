import openpyxl

def convert_sq_meter_to_sq_km(value):
    return value / 1_000_000

def process_excel(file_path):
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(r'D:\Working\Working2\New_Output\GEEimages\Bagmati_march\ExcelSheet\DroughtArea.xlsx')
    ws = wb.active

    # Iterate over all cells in the sheet
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            try:
                # Convert the value from sq meters to sq kilometers
                cell.value = convert_sq_meter_to_sq_km(float(cell.value))
            except (TypeError, ValueError):
                # If the cell value is not a number, skip it
                continue

    # Save the modified workbook
    wb.save(file_path)

# Path to your Excel file
file_path = r'D:\Working\Working2\New_Output\GEEimages\Bagmati_march\ExcelSheet\DroughtAreaKM.xlsx'

process_excel(file_path)
